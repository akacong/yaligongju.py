import sys

import openpyxl
import pandas as pd
from PyQt5 import Qt, QtCore
from PyQt5.QtGui import QStandardItem, QStandardItemModel
from PyQt5.QtWidgets import QMainWindow, QFileDialog, QApplication

from YALIGONGJU import Ui_MainWindow


def calP0(r, h, g):
    p0 = r * g * h
    return p0 / 1000


def calTTI(rm, t):
    tti = rm * t
    return tti


def calPh(h, g):
    ph = g * h
    return ph / 1000


def calPp(p0, h, alf, tn_t, n):
    pn = 0.00981 / h
    pp = p0 - (p0 - pn) * alf * tn_t ** n
    return pp


def calPE(p0, pp):
    pe = p0 - pp
    return pe


def calTTPI(tm, rm, pe):
    ttpi = tm * rm * pe
    return ttpi


class win2(Ui_MainWindow, QMainWindow):
    def __init__(self):
        super(win2, self).__init__()
        self.setupUi(self)
        self.modle = QStandardItemModel()
        self.modle2 = QStandardItemModel()
        self.modle3 = QStandardItemModel()
        self.modle4 = QStandardItemModel()
        self.modle5 = QStandardItemModel()
        self.modle6 = QStandardItemModel()
        self.pushButton.clicked.connect(self.openfile2)
        self.pushButton_2.clicked.connect(self.calculate)
        self.pushButton_3.clicked.connect(self.select2)
        self.pushButton_4.clicked.connect(self.savefile)

    def openfile(self):
        self.listWidget.clear()
        self.filePath, filetype = QFileDialog.getOpenFileName(self, "选取文件", "./",
                                                              "All Files (*);;Text Files (*.txt)")
        self.lineEdit.setText(self.filePath)
        try:
            self.rdfile = openpyxl.load_workbook(self.filePath)
            self.tablename = self.rdfile.sheetnames
            self.listWidget.addItems(self.tablename)
        except Exception as e:
            print(e)

    def select(self):
        i = int(self.spinBox.text())
        if i == 1:
            sheetname = self.listWidget.currentItem().text()
            df = pd.DataFrame(pd.read_excel(self.filePath, sheet_name=sheetname))
            data = df.values
            colms = df.columns.tolist()
            for idx in range(len(colms)):
                self.modle.setHorizontalHeaderItem(idx, QStandardItem(colms[idx]))
            row = 0
            for line in data:
                col = 0
                for item in line:
                    self.modle.setItem(row, col, QStandardItem(str(item)))
                    col += 1
                row += 1
            self.tableView.setModel(self.modle)
        elif i == 2:
            sheetname = self.listWidget.currentItem().text()
            df = pd.DataFrame(pd.read_excel(self.filePath, sheet_name=sheetname))
            data = df.values
            colms = df.columns.tolist()
            for idx in range(len(colms)):
                self.modle3.setHorizontalHeaderItem(idx, QStandardItem(colms[idx]))
            row = 0
            for line in data:
                col = 0
                for item in line:
                    self.modle3.setItem(row, col, QStandardItem(str(item)))
                    col += 1
                row += 1
            self.tableView_3.setModel(self.modle3)
        else:
            sheetname = self.listWidget.currentItem().text()
            df = pd.DataFrame(pd.read_excel(self.filePath, sheet_name=sheetname))
            data = df.values
            colms = df.columns.tolist()
            for idx in range(len(colms)):
                self.modle5.setHorizontalHeaderItem(idx, QStandardItem(colms[idx]))
            row = 0
            for line in data:
                col = 0
                for item in line:
                    self.modle5.setItem(row, col, QStandardItem(str(item)))
                    col += 1
                row += 1
            self.tableView_4.setModel(self.modle5)

    def openfile2(self):
        self.listWidget.clear()
        file_dialog = QFileDialog()
        file_dialog.setFileMode(QFileDialog.ExistingFiles)
        file_dialog.setNameFilter("All Files (*.*)")  # 设置文件过滤器
        if file_dialog.exec_():
            file_paths = file_dialog.selectedFiles()
            for file_path in file_paths:
                self.rdfile = openpyxl.load_workbook(file_path)
                self.tablename = self.rdfile.sheetnames
                for item in self.tablename:
                    self.listWidget.addItem(file_path + "//" + item)

    def select2(self):
        i = int(self.spinBox.text())
        if i == 1:
            current_item = self.listWidget.currentItem()
            filepath_sheetname = current_item.text()
            filepathAndSheetname = filepath_sheetname.split("//")
            filepath = filepathAndSheetname[0]
            sheetname = filepathAndSheetname[1]
            df = pd.DataFrame(pd.read_excel(filepath, sheet_name=sheetname))
            data = df.values
            colms = df.columns.tolist()
            for idx in range(len(colms)):
                self.modle.setHorizontalHeaderItem(idx, QStandardItem(colms[idx]))
            row = 0
            for line in data:
                col = 0
                for item in line:
                    self.modle.setItem(row, col, QStandardItem(str(item)))
                    col += 1
                row += 1
            self.tableView.setModel(self.modle)
        elif i == 2:
            current_item = self.listWidget.currentItem()
            filepath_sheetname = current_item.text()
            filepathAndSheetname = filepath_sheetname.split("//")
            filepath = filepathAndSheetname[0]
            sheetname = filepathAndSheetname[1]
            df = pd.DataFrame(pd.read_excel(filepath, sheet_name=sheetname))
            data = df.values
            colms = df.columns.tolist()
            for idx in range(len(colms)):
                self.modle3.setHorizontalHeaderItem(idx, QStandardItem(colms[idx]))
            row = 0
            for line in data:
                col = 0
                for item in line:
                    self.modle3.setItem(row, col, QStandardItem(str(item)))
                    col += 1
                row += 1
            self.tableView_3.setModel(self.modle3)
        else:
            current_item = self.listWidget.currentItem()
            filepath_sheetname = current_item.text()
            filepathAndSheetname = filepath_sheetname.split("//")
            filepath = filepathAndSheetname[0]
            sheetname = filepathAndSheetname[1]
            df = pd.DataFrame(pd.read_excel(filepath, sheet_name=sheetname))
            data = df.values
            colms = df.columns.tolist()
            for idx in range(len(colms)):
                self.modle5.setHorizontalHeaderItem(idx, QStandardItem(colms[idx]))
            row = 0
            for line in data:
                col = 0
                for item in line:
                    self.modle5.setItem(row, col, QStandardItem(str(item)))
                    col += 1
                row += 1
            self.tableView_4.setModel(self.modle5)

    def calculate(self):
        g = float(self.lineEdit_4.text())
        alf = float(self.lineEdit_5.text())
        n = float(self.lineEdit_6.text())
        # page1
        rows = self.tableView.model().rowCount()
        cols = self.tableView.model().columnCount()
        df1 = pd.DataFrame(columns=[self.tableView.model().headerData(i, QtCore.Qt.Horizontal) for i in range(cols)])
        for row in range(rows):
            data = []
            for col in range(cols):
                if col == 0:
                    index = self.tableView.model().index(row, col)
                    data.append(self.tableView.model().data(index))
                else:
                    index = self.tableView.model().index(row, col)
                    value = float(self.tableView.model().data(index))
                    data.append(value)
            df1.loc[len(df1)] = data
        r_m1, time1, depth1, rho1, tn_t1 = df1['2^m'], df1['Δtm时间差（Ma）'], df1['Δtm时间埋深（m）'], df1['rho（g/cm3）'], df1['Δtn/Δt']
        tti1 = calTTI(r_m1, time1)
        p01 = round(calP0(rho1, depth1, g), 2)
        ph1 = round(calPh(depth1, g), 2)
        pp1 = round(calPp(p01, depth1, alf, tn_t1, n), 2)
        pe1 = round(calPE(p01, pp1), 2)
        ttpi1 = round(calTTPI(time1, r_m1, pe1), 2)
        sig_tti = round(sum(tti1), 2)
        sig_ttpi = round(sum(ttpi1), 2)
        colms = ['△TTI ', 'P0m（Mpa）', ' Phm（Mpa）', 'Ppm（Mpa）', 'Pe（Mpa）', '△TTPI']
        outputs = pd.concat([tti1, p01, ph1, pp1, pe1, ttpi1], axis=1, ignore_index=True)
        outputs.columns = colms
        data = outputs.values
        cols = outputs.columns.tolist()
        for idx in range(len(cols)):
            self.modle2.setHorizontalHeaderItem(idx, QStandardItem(cols[idx]))
        row = 0
        for line in data:
            col = 0
            for item in line:
                self.modle2.setItem(row, col, QStandardItem(str(item)))
                col += 1
            row += 1
        self.tableView_2.setModel(self.modle2)
        self.lineEdit_2.setText(str(sig_tti))
        self.lineEdit_3.setText(str(sig_ttpi))
        # page2
        rows = self.tableView_3.model().rowCount()
        cols = self.tableView_3.model().columnCount()
        df2 = pd.DataFrame(columns=[self.tableView_3.model().headerData(i, QtCore.Qt.Horizontal) for i in range(cols)])
        for row in range(rows):
            data = []
            for col in range(cols):
                if col == 0:
                    index = self.tableView_3.model().index(row, col)
                    data.append(self.tableView_3.model().data(index))
                else:
                    index = self.tableView_3.model().index(row, col)
                    value = float(self.tableView_3.model().data(index))
                    data.append(value)
            df2.loc[len(df2)] = data
        r_m2, time2, depth2, rho2, tn_t2 = df2['2^m'], df2['Δtm时间差（Ma）'], df2['Δtm时间埋深（m）'], df2['rho（g/cm3）'], df2['Δtn/Δt']
        tti2 = calTTI(r_m2, time2)
        p02 = round(calP0(rho2, depth2, g), 2)
        ph2 = round(calPh(depth2, g), 2)
        pp2 = round(calPp(p02, depth2, alf, tn_t2, n), 2)
        pe2 = round(calPE(p02, pp2), 2)
        ttpi2 = round(calTTPI(time2, r_m2, pe2), 2)
        sig_tti2 = round(sum(tti2), 2)
        sig_ttpi2 = round(sum(ttpi2), 2)
        colms = ['△TTI ', 'P0m（Mpa）', ' Phm（Mpa）', 'Ppm（Mpa）', 'Pe（Mpa）', '△TTPI']
        outputs = pd.concat([tti2, p02, ph2, pp2, pe2, ttpi2], axis=1, ignore_index=True)
        outputs.columns = colms
        data = outputs.values
        cols = outputs.columns.tolist()
        for idx in range(len(cols)):
            self.modle4.setHorizontalHeaderItem(idx, QStandardItem(cols[idx]))
        row = 0
        for line in data:
            col = 0
            for item in line:
                self.modle4.setItem(row, col, QStandardItem(str(item)))
                col += 1
            row += 1
        self.tableView_5.setModel(self.modle4)
        self.lineEdit_7.setText(str(sig_tti2))
        self.lineEdit_8.setText(str(sig_ttpi2))
        # page3
        rows = self.tableView_4.model().rowCount()
        cols = self.tableView_4.model().columnCount()
        df3 = pd.DataFrame(columns=[self.tableView_4.model().headerData(i, QtCore.Qt.Horizontal) for i in range(cols)])
        for row in range(rows):
            data = []
            for col in range(cols):
                if col == 0:
                    index = self.tableView_4.model().index(row, col)
                    data.append(self.tableView_4.model().data(index))
                else:
                    index = self.tableView_4.model().index(row, col)
                    value = float(self.tableView_4.model().data(index))
                    data.append(value)
            df3.loc[len(df3)] = data
        r_m3, time3, depth3, rho3, tn_t3 = df3['2^m'], df3['Δtm时间差（Ma）'], df3['Δtm时间埋深（m）'], df3['rho（g/cm3）'], df3['Δtn/Δt']
        tti3 = calTTI(r_m3, time3)
        p03 = round(calP0(rho3, depth3, g), 2)
        ph3 = round(calPh(depth3, g), 2)
        pp3 = round(calPp(p03, depth3, alf, tn_t3, n), 2)
        pe3 = round(calPE(p03, pp3), 2)
        ttpi3 = round(calTTPI(time3, r_m3, pe3), 2)
        sig_tti3 = round(sum(tti3), 2)
        sig_ttpi3 = round(sum(ttpi3), 2)
        colms = ['△TTI ', 'P0m（Mpa）', ' Phm（Mpa）', 'Ppm（Mpa）', 'Pe（Mpa）', '△TTPI']
        outputs = pd.concat([tti3, p03, ph3, pp3, pe3, ttpi3], axis=1, ignore_index=True)
        outputs.columns = colms
        data = outputs.values
        cols = outputs.columns.tolist()
        for idx in range(len(cols)):
            self.modle6.setHorizontalHeaderItem(idx, QStandardItem(cols[idx]))
        row = 0
        for line in data:
            col = 0
            for item in line:
                self.modle6.setItem(row, col, QStandardItem(str(item)))
                col += 1
            row += 1
        self.tableView_6.setModel(self.modle6)
        self.lineEdit_9.setText(str(sig_tti3))
        self.lineEdit_10.setText(str(sig_ttpi3))

    def savefile(self):
        rows = self.tableView_2.model().rowCount()
        cols = self.tableView_2.model().columnCount()
        df1 = pd.DataFrame(columns=[self.tableView_2.model().headerData(i, QtCore.Qt.Horizontal) for i in range(cols)])
        for row in range(rows):
            data = []
            for col in range(cols):
                index = self.tableView_2.model().index(row, col)
                value = float(self.tableView_2.model().data(index))
                data.append(value)
            df1.loc[len(df1)] = data
        rows = self.tableView_5.model().rowCount()
        cols = self.tableView_5.model().columnCount()
        df2 = pd.DataFrame(columns=[self.tableView_5.model().headerData(i, QtCore.Qt.Horizontal) for i in range(cols)])
        for row in range(rows):
            data = []
            for col in range(cols):
                index = self.tableView_5.model().index(row, col)
                value = float(self.tableView_5.model().data(index))
                data.append(value)
            df2.loc[len(df2)] = data
        rows = self.tableView_6.model().rowCount()
        cols = self.tableView_6.model().columnCount()
        df3 = pd.DataFrame(columns=[self.tableView_6.model().headerData(i, QtCore.Qt.Horizontal) for i in range(cols)])
        for row in range(rows):
            data = []
            for col in range(cols):
                index = self.tableView_6.model().index(row, col)
                value = float(self.tableView_6.model().data(index))
                data.append(value)
            df3.loc[len(df3)] = data
        filename, filetype = QFileDialog.getSaveFileName(self, "Save File", "output.xlsx", "Excel files (*.xlsx)")
        if filename:
            writer = pd.ExcelWriter(filename)
            df1.to_excel(writer, sheet_name="output1", index=False)
            df2.to_excel(writer, sheet_name="output2", index=False)
            df3.to_excel(writer, sheet_name="output3", index=False)
            #writer.save()
            writer._save()

if __name__ == '__main__':
    app = QApplication(sys.argv)
    mywindow = win2()
    mywindow.show()
    sys.exit(app.exec_())
